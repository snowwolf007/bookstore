"""书刊 + 分类管理 API"""
import os
import json
import uuid
import shutil
from pathlib import Path
from typing import Optional, List

from fastapi import (
    APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
)
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from ..config import settings
from ..database import get_db
from ..models import Book, Category
from ..schemas import (
    BookCreate, BookUpdate, BookOut,
    CategoryCreate, CategoryOut, CategoryTree,
)
from .auth import get_current_user

router = APIRouter(prefix="/api/admin", tags=["管理后台"])
public_router = APIRouter(prefix="/api/books", tags=["公开接口"])

ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp"}


# ========== 文件上传 ==========
def save_upload(file: UploadFile, subdir: str = "covers") -> str:
    """保存上传文件，返回相对路径"""
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(400, f"不支持的图片格式: {ext}")
    # 生成唯一文件名
    filename = f"{uuid.uuid4().hex}{ext}"
    save_dir = settings.UPLOAD_DIR / subdir
    save_dir.mkdir(parents=True, exist_ok=True)
    save_path = save_dir / filename
    with open(save_path, "wb") as f:
        f.write(file.file.read())
    return f"uploads/{subdir}/{filename}"


# ========== 分类管理 ==========
@router.get("/categories", response_model=List[CategoryTree])
async def list_categories(db: AsyncSession = Depends(get_db)):
    """获取分类树"""
    result = await db.execute(select(Category).order_by(Category.sort_order))
    categories = result.scalars().all()
    # 构建树
    cat_map = {c.id: CategoryTree(
        id=c.id, name=c.name, slug=c.slug,
        parent_id=c.parent_id, sort_order=c.sort_order,
        is_active=c.is_active, children=[]
    ) for c in categories}
    tree = []
    for c in cat_map.values():
        if c.parent_id and c.parent_id in cat_map:
            cat_map[c.parent_id].children.append(c)
        else:
            tree.append(c)
    return tree


@router.post("/categories", response_model=CategoryOut)
async def create_category(data: CategoryCreate, db: AsyncSession = Depends(get_db)):
    cat = Category(**data.model_dump())
    db.add(cat)
    await db.commit()
    await db.refresh(cat)
    return cat


@router.put("/categories/{cat_id}", response_model=CategoryOut)
async def update_category(cat_id: int, data: CategoryCreate,
                          db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "分类不存在")
    for k, v in data.model_dump().items():
        setattr(cat, k, v)
    await db.commit()
    await db.refresh(cat)
    return cat


@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "分类不存在")
    await db.delete(cat)
    await db.commit()
    return {"ok": True}


# ========== 书刊管理 ==========
@router.get("/books", response_model=List[dict])
async def list_books(
    category_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    published: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """管理后台：书刊列表（含分页、搜索、筛选）"""
    query = select(Book).options(joinedload(Book.category)).order_by(Book.sort_order, Book.id.desc())

    if category_id:
        query = query.where(Book.category_id == category_id)
    if published is not None:
        query = query.where(Book.is_published == published)
    if search:
        like = f"%{search}%"
        query = query.where(
            or_(Book.title.like(like), Book.author.like(like),
                Book.isbn.like(like), Book.publisher.like(like))
        )
    if page_size > 0:
        query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    books = result.unique().scalars().all()
    return [b.to_dict() for b in books]


@router.get("/books/count")
async def count_books(
    search: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    query = select(func.count(Book.id))
    if search:
        like = f"%{search}%"
        query = query.where(
            or_(Book.title.like(like), Book.author.like(like), Book.isbn.like(like))
        )
    if category_id:
        query = query.where(Book.category_id == category_id)
    result = await db.execute(query)
    return {"count": result.scalar()}


@router.post("/books", response_model=BookOut)
async def create_book(
    title: str = Form(...),
    title_en: str = Form(""),
    title_cn: str = Form(""),
    author: str = Form(""),
    publisher: str = Form(""),
    isbn: str = Form(""),
    year: str = Form(""),
    price: float = Form(0),
    stock: int = Form(0),
    description: str = Form(""),
    catalog_text: str = Form(""),
    category_id: int = Form(0),
    tags: str = Form("[]"),  # JSON字符串
    is_published: bool = Form(True),
    is_featured: bool = Form(False),
    cover_path: str = Form(""),
    cover: UploadFile = File(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    # 使用表单传入的 cover_path，或从上传文件获取
    if cover:
        cover_path_final = save_upload(cover, "covers")
    else:
        cover_path_final = cover_path

    tags_list = json.loads(tags) if tags else []
    cat_id = category_id if category_id > 0 else None

    book = Book(
        title=title, title_en=title_en, title_cn=title_cn,
        author=author, publisher=publisher, isbn=isbn, year=year,
        price=price, stock=stock, description=description,
        catalog_text=catalog_text, cover_path=cover_path_final,
        category_id=cat_id, tags=tags_list,
        is_published=is_published, is_featured=is_featured,
    )
    db.add(book)
    await db.commit()
    # 重新查询以加载关联关系
    result = await db.execute(
        select(Book).options(joinedload(Book.category)).where(Book.id == book.id)
    )
    book = result.scalar_one()
    return book.to_dict()


@router.put("/books/{book_id}", response_model=BookOut)
async def update_book(
    book_id: int,
    title: str = Form(None),
    title_en: str = Form(None),
    title_cn: str = Form(None),
    author: str = Form(None),
    publisher: str = Form(None),
    isbn: str = Form(None),
    year: str = Form(None),
    price: float = Form(None),
    stock: int = Form(None),
    description: str = Form(None),
    catalog_text: str = Form(None),
    category_id: int = Form(None),
    tags: str = Form(None),
    is_published: bool = Form(None),
    is_featured: bool = Form(None),
    cover_path: str = Form(None),
    cover: UploadFile = File(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(404, "书刊不存在")

    fields = {
        "title": title, "title_en": title_en, "title_cn": title_cn,
        "author": author, "publisher": publisher, "isbn": isbn,
        "year": year, "price": price, "stock": stock,
        "description": description, "catalog_text": catalog_text,
        "is_published": is_published, "is_featured": is_featured,
        "cover_path": cover_path,
    }
    for k, v in fields.items():
        if v is not None:
            setattr(book, k, v)

    if category_id is not None:
        book.category_id = category_id if category_id > 0 else None
    if tags is not None:
        book.tags = json.loads(tags)
    if cover:
        book.cover_path = save_upload(cover, "covers")

    await db.commit()
    # 重新查询以加载关联关系
    result = await db.execute(
        select(Book).options(joinedload(Book.category)).where(Book.id == book.id)
    )
    book = result.scalar_one()
    return book.to_dict()


@router.delete("/books/{book_id}")
async def delete_book(book_id: int, db: AsyncSession = Depends(get_db),
                      _=Depends(get_current_user)):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(404, "书刊不存在")
    await db.delete(book)
    await db.commit()
    return {"ok": True}


@router.post("/books/{book_id}/catalog-images")
async def upload_catalog_image(
    book_id: int,
    image: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """上传目录页图片"""
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(404, "书刊不存在")
    path = save_upload(image, "catalogs")
    imgs = book.catalog_images or []
    imgs.append(path)
    book.catalog_images = imgs
    await db.commit()
    return {"path": path, "images": imgs}


# ========== 公开接口（小程序/前端使用） ==========
@public_router.get("/")
async def public_list_books(
    category_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    featured: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
):
    """公开：书刊列表（仅已上架）"""
    query = select(Book).options(joinedload(Book.category)).where(
        Book.is_published == True
    ).order_by(Book.sort_order, Book.id.desc())

    if category_id:
        query = query.where(Book.category_id == category_id)
    if featured:
        query = query.where(Book.is_featured == True)
    if search:
        like = f"%{search}%"
        query = query.where(
            or_(Book.title.like(like), Book.author.like(like),
                Book.isbn.like(like), Book.publisher.like(like))
        )

    total_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(total_q)).scalar()

    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    books = result.unique().scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [b.to_dict() for b in books],
    }


@public_router.get("/categories")
async def public_categories(db: AsyncSession = Depends(get_db)):
    """公开：分类列表（仅含已上架书籍的分类）"""
    result = await db.execute(
        select(Category).where(Category.is_active == True).order_by(Category.sort_order)
    )
    return [c.to_dict() for c in result.scalars().all()]


@public_router.get("/{book_id}")
async def public_book_detail(book_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Book).options(joinedload(Book.category)).where(Book.id == book_id)
    )
    book = result.scalar_one_or_none()
    if not book or not book.is_published:
        raise HTTPException(404, "书刊不存在")
    return book.to_dict()


# ========== Excel 批量导入 ==========
import re as _re
import uuid as _uuid


@router.post("/books/excel-import")
async def import_books_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """从 Excel 文件批量导入书刊"""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的 Excel 文件")

    # 保存上传的 Excel
    ext = Path(file.filename).suffix.lower()
    tmp_path = settings.UPLOAD_DIR / f"_import_{_uuid.uuid4().hex}{ext}"
    tmp_path.parent.mkdir(parents=True, exist_ok=True)
    with open(tmp_path, "wb") as f:
        f.write(file.file.read())

    try:
        import openpyxl
    except ImportError:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(500, "服务器缺少 openpyxl 库，请联系管理员安装")

    try:
        wb = openpyxl.load_workbook(tmp_path)
    except Exception as e:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, f"无法读取 Excel 文件: {e}")

    ws = wb.active

    # 解析列名
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # 列名映射
    col_map = {}
    for i, h in enumerate(headers, 1):
        h_clean = h.strip()
        if h_clean == "书名（必填）" or h_clean == "书名":
            col_map[i] = "title"
        elif h_clean == "英文书名":
            col_map[i] = "title_en"
        elif h_clean == "中文书名":
            col_map[i] = "title_cn"
        elif h_clean == "作者":
            col_map[i] = "author"
        elif h_clean == "出版社":
            col_map[i] = "publisher"
        elif h_clean == "ISBN":
            col_map[i] = "isbn"
        elif h_clean == "出版年份":
            col_map[i] = "year"
        elif h_clean == "定价":
            col_map[i] = "price"
        elif h_clean == "库存":
            col_map[i] = "stock"
        elif h_clean == "分类":
            col_map[i] = "category_name"
        elif h_clean == "标签":
            col_map[i] = "tags"
        elif h_clean == "简介":
            col_map[i] = "description"
        elif h_clean == "目录":
            col_map[i] = "catalog_text"
        elif h_clean == "封面文件名":
            col_map[i] = "cover_path"
        elif h_clean == "上架":
            col_map[i] = "is_published"

    if "title" not in col_map.values():
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, "Excel 缺少「书名」列，请使用模板格式")

    # 读取数据行
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        vals = [v for v in row if v is not None and str(v).strip()]
        if not vals:
            continue
        book = {}
        for col_idx in col_map:
            val = row[col_idx - 1]
            if val is not None:
                book[col_map[col_idx]] = str(val).strip()
            else:
                book[col_map[col_idx]] = ""
        if book.get("title", "").strip():
            rows.append(book)

    if not rows:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, "Excel 中没有有效数据")

    # 获取已有分类映射
    result = await db.execute(select(Category))
    all_cats = result.scalars().all()
    cat_map = {}
    for c in all_cats:
        key = _re.sub(r"[\s_\-:：，,\.\(\)（）\[\]\/#&]", "", c.name).lower()
        cat_map[key] = c.id

    # 逐条导入
    imported = 0
    skipped = 0
    errors = []

    for i, book_data in enumerate(rows):
        title = book_data.get("title", "").strip()
        if not title:
            continue

        # 去重
        q = await db.execute(select(Book).where(Book.title == title))
        if q.scalar_one_or_none():
            skipped += 1
            continue

        # 分类
        cat_name = book_data.get("category_name", "")
        category_id = None
        if cat_name:
            key = _re.sub(r"[\s_\-:：，,\.\(\)（）\[\]\/#&]", "", cat_name).lower()
            category_id = cat_map.get(key)

        # 标签
        tags_str = book_data.get("tags", "")
        tags = [t.strip() for t in _re.split(r"[,，/、]", tags_str) if t.strip()]

        # 上架状态
        pub_str = book_data.get("is_published", "是")
        is_published = pub_str.lower() in ("是", "yes", "true", "1", "y")

        # 封面路径
        cover = book_data.get("cover_path", "")
        if cover:
            cover = cover.replace("\\", "/")
            if not cover.startswith("uploads/"):
                cover = f"uploads/{cover}"

        # 数值
        try:
            price = float(book_data.get("price", "0"))
        except ValueError:
            price = 0
        try:
            stock = int(float(book_data.get("stock", "1")))
        except ValueError:
            stock = 1

        book = Book(
            title=title,
            title_en=book_data.get("title_en", ""),
            title_cn=book_data.get("title_cn", ""),
            author=book_data.get("author", ""),
            publisher=book_data.get("publisher", ""),
            isbn=book_data.get("isbn", ""),
            year=book_data.get("year", ""),
            price=price,
            stock=stock,
            description=book_data.get("description", ""),
            catalog_text=book_data.get("catalog_text", ""),
            cover_path=cover,
            category_id=category_id,
            tags=tags,
            is_published=is_published,
            sort_order=i + 1,
        )
        db.add(book)
        imported += 1

        # 每 50 条提交一次
        if imported % 50 == 0:
            await db.commit()

    await db.commit()
    tmp_path.unlink(missing_ok=True)

    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "total": len(rows),
        "errors": errors[:10],
    }


# ========== 自动搜索封面 ==========

@router.post("/books/{book_id}/fetch-cover")
async def fetch_book_cover(
    book_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """为单本书自动搜索并下载封面（OpenLibrary → Google Books）"""
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(404, "书刊不存在")

    from ..services.cover_fetcher import fetch_and_save_cover

    try:
        cover_path = await fetch_and_save_cover(book)
        if cover_path:
            book.cover_path = cover_path
            await db.commit()
            return {"success": True, "cover_path": cover_path}
        else:
            return {"success": False, "message": "未找到封面"}
    except Exception as e:
        raise HTTPException(500, f"搜索封面失败: {e}")


@router.post("/books/batch/cover-fetch")
async def batch_fetch_covers(
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """为所有无封面书刊并行批量搜索封面（asyncio.gather，大幅提速）"""
    import asyncio

    result = await db.execute(
        select(Book).where(
            (Book.cover_path == "") | (Book.cover_path.is_(None))
        ).order_by(Book.id)
    )
    books = result.scalars().all()

    if not books:
        return {"success": True, "total": 0, "fetched": 0, "message": "全部已有封面"}

    from ..services.cover_fetcher import fetch_and_save_cover

    async def search_one(book) -> dict:
        """搜索单本书封面"""
        try:
            cover_path = await asyncio.wait_for(
                fetch_and_save_cover(book), timeout=15
            )
            if cover_path:
                book.cover_path = cover_path
                return {"id": book.id, "title": book.title[:30], "status": "ok"}
            else:
                return {"id": book.id, "title": book.title[:30], "status": "not_found"}
        except asyncio.TimeoutError:
            return {"id": book.id, "title": book.title[:30], "status": "timeout"}
        except Exception as e:
            return {"id": book.id, "title": book.title[:30], "status": f"error: {e}"}

    # 并行搜索（每次5本并发，避免被反爬封IP也不要撑爆连接池）
    results = []
    batch_size = 5
    for i in range(0, len(books), batch_size):
        batch = books[i:i + batch_size]
        batch_results = await asyncio.gather(*[search_one(b) for b in batch])
        results.extend(batch_results)

        # 每批提交一次
        for br in batch_results:
            if br["status"] == "ok":
                await db.commit()

    await db.commit()
    fetched = sum(1 for r in results if r["status"] == "ok")

    return {
        "success": True,
        "total": len(books),
        "fetched": fetched,
        "results": results,
    }
